from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from bs4 import BeautifulSoup
import requests
import re
from datetime import datetime
import sqlite3
import io
#from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

app = Flask(__name__, static_folder='static', template_folder='static')
CORS(app)

DB_PATH = 'data/bills.db'

def init_db():
    os.makedirs('data', exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS bills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            comercializadora TEXT NOT NULL,
            periodo TEXT,
            tarifa TEXT,
            precio REAL,
            energia_verde BOOLEAN,
            revision TEXT,
            permanencia TEXT,
            servicios TEXT,
            tipo TEXT,
            url TEXT,
            fecha_captura TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def add_bill(data):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        INSERT INTO bills (comercializadora, periodo, tarifa, precio, energia_verde, revision, permanencia, servicios, tipo, url)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (data['comercializadora'], data['periodo'], data['tarifa'], data['precio'], data['energia_verde'], data['revision'], data['permanencia'], data['servicios'], data['tipo'], data['url']))
    conn.commit()
    bill_id = c.lastrowid
    conn.close()
    return bill_id

def get_all_bills():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT id, comercializadora, periodo, tarifa, precio, energia_verde, revision, permanencia, servicios, tipo, url, fecha_captura FROM bills ORDER BY fecha_captura DESC')
    bills = c.fetchall()
    conn.close()
    return bills

def delete_bill(bill_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('DELETE FROM bills WHERE id = ?', (bill_id,))
    conn.commit()
    conn.close()

def clear_all_bills():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('DELETE FROM bills')
    conn.commit()
    conn.close()

def scrape_cnmc_url(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        response = requests.get(url, headers=headers, timeout=10)
        response.encoding = 'utf-8'
        
        if response.status_code != 200:
            return {'error': f'No se pudo acceder a la URL (código {response.status_code})'}
        
        soup = BeautifulSoup(response.content, 'html.parser')
        all_text = soup.get_text()
        
        comercializadora_match = re.search(r'Nombre comercializadora[:\s]*([^\n<]+)', all_text)
        periodo_match = re.search(r'Periodo facturación[:\s]*(\d{2}/\d{2}/\d{4}\s*-\s*\d{2}/\d{2}/\d{4})', all_text)
        precio_match = re.search(r'(\d+[.,]\d{2})\s*€', all_text)
        
        bill_data = {
            'comercializadora': comercializadora_match.group(1).strip() if comercializadora_match else 'No identificada',
            'periodo': periodo_match.group(1).strip() if periodo_match else 'N/A',
            'tarifa': 'Tarifa con 3 precios fijos',
            'precio': float(precio_match.group(1).replace(',', '.')) if precio_match else 0.0,
            'energia_verde': True,
            'revision': 'Revisión anual',
            'permanencia': 'Sin permanencia',
            'servicios': 'Sin servicios adicionales',
            'tipo': 'current',
            'url': url
        }
        return bill_data
    except Exception as e:
        return {'error': f'Error al procesar: {str(e)}'}

@app.route('/api/scrape', methods=['POST'])
def scrape():
    data = request.json
    url = data.get('url', '')
    
    if not url:
        return jsonify({'error': 'URL no proporcionada'}), 400
    
    result = scrape_cnmc_url(url)
    
    if 'error' in result:
        return jsonify(result), 400
    
    bill_id = add_bill(result)
    result['id'] = bill_id
    
    return jsonify(result), 200

@app.route('/api/bills', methods=['GET'])
def get_bills():
    bills = get_all_bills()
    bills_list = []
    for bill in bills:
        bills_list.append({
            'id': bill[0], 'comercializadora': bill[1], 'periodo': bill[2], 'tarifa': bill[3],
            'precio': bill[4], 'energia_verde': bill[5], 'revision': bill[6],
            'permanencia': bill[7], 'servicios': bill[8], 'tipo': bill[9], 'url': bill[10], 'fecha_captura': bill[11]
        })
    return jsonify(bills_list), 200

@app.route('/api/bills/<int:bill_id>', methods=['DELETE'])
def delete_bill_route(bill_id):
    delete_bill(bill_id)
    return jsonify({'success': True}), 200

@app.route('/api/bills', methods=['DELETE'])
def clear_bills():
    clear_all_bills()
    return jsonify({'success': True}), 200

@app.route('/api/download/excel', methods=['GET'])
def download_excel():
    bills = get_all_bills()
    if not bills:
        return jsonify({'error': 'No hay datos para descargar'}), 400
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"
    
    header_fill = PatternFill(start_color="2180a3", end_color="2180a3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ['ID', 'Comercializadora', 'Período', 'Tipo Tarifa', 'Precio (€)', 'Energía Verde', 
               'Revisión', 'Permanencia', 'Servicios', 'Tipo', 'URL', 'Fecha Captura']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row, bill in enumerate(bills, 2):
        ws.cell(row=row, column=1).value = bill[0]
        ws.cell(row=row, column=2).value = bill[1]
        ws.cell(row=row, column=3).value = bill[2]
        ws.cell(row=row, column=4).value = bill[3]
        ws.cell(row=row, column=5).value = bill[4]
        ws.cell(row=row, column=6).value = "Sí" if bill[5] else "No"
        ws.cell(row=row, column=7).value = bill[6]
        ws.cell(row=row, column=8).value = bill[7]
        ws.cell(row=row, column=9).value = bill[8]
        ws.cell(row=row, column=10).value = bill[9]
        ws.cell(row=row, column=11).value = bill[10]
        ws.cell(row=row, column=12).value = bill[11]
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'facturas-energia-{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/')
def index():
    return send_file('static/app_publica.html')

@app.route('/admin')
def admin():
    return send_file('static/admin_panel.html')

if __name__ == '__main__':
    init_db()
    print("🚀 Servidor Flask iniciado en http://localhost:5000")
    print("📱 App pública: http://localhost:5000")
    print("⚙️ Panel admin: http://localhost:5000/admin")
    app.run(debug=True, host='0.0.0.0', port=5000)
